# utils/data_handler.py
import os
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from fuzzywuzzy import fuzz
from config import EXCEL_PATH

# استيراد speak عشان ينطق بعد الحفظ
from utils.voice_engine import speak

def parse_time(time_raw):
    if not time_raw:
        return "not_found"
    time_raw = time_raw.lower()
    table = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')
    time_raw = time_raw.translate(table)
    replacements = {
        "ونصف": "30", "ونص": "30", "وربع": "15",
        "وثلث": "20", "وتلت": "20", "دقيقة": "", "ساعة": ""
    }
    for word, replacement in replacements.items():
        if word in time_raw:
            time_raw = time_raw.replace(word, replacement)
    numbers = re.findall(r'\d+', time_raw)
    hour, minute = 0, 0
    if len(numbers) >= 2:
        hour, minute = int(numbers[0]), int(numbers[1])
    elif len(numbers) == 1:
        hour, minute = int(numbers[0]), 0
    else:
        return "not_found"
    if 1 <= hour <= 12 and 0 <= minute <= 59:
        return f"{hour}:{minute:02d}"
    else:
        return "invalid"

# 🔴 التعديل هنا: إضافة appointment_day
def save_appointment_details(file_path, appointment_type, appointment_time, am_pm, appointment_day):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Type", "Time", "Period", "Day"]) # 🔴 إضافة عمود اليوم
        wb.save(file_path)
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        ws.append([appointment_type, appointment_time, am_pm, appointment_day]) # 🔴 حفظ اليوم
        wb.save(file_path)
        
        # 🔴 نطق التأكيد صوتياً شامل اليوم
        speak(f"تم إضافة الموعد: {appointment_type} {appointment_day} في تمام الساعة {appointment_time} {am_pm}")
        print(f"✅ Saved & Confirmed: {appointment_type} | {appointment_day} at {appointment_time} {am_pm}")
        
    except Exception as e:
        print(f"❌ Error saving file: {e}")
        speak("حدث خطأ أثناء حفظ الموعد")

def display_appointments():
    try:
        # قراءة الملف
        df = pd.read_excel(EXCEL_PATH)
        
        # 1. تنظيف البيانات: حذف الصفوف اللي كلها فاضية تماماً (الـ NaN اللي ظاهرة عندك)
        df = df.dropna(how='all').reset_index(drop=True)
        
        # 2. التأكد من أسماء الأعمدة: لو الإكسل مش قاري العناوين صح، بنسميها يدوي
        if 'Period' not in df.columns:
            # 🔴 تعديل قراءة الأعمدة لتشمل اليوم
            df.columns = ['Type', 'Time', 'Period', 'Day'] + list(df.columns[4:])

        print("\n--- Scheduled Appointments ---")
        print(df.to_string(index=False))
        
        if not df.empty:
            schedule_text = "مواعيدك الحالية هي: "
            for _, row in df.iterrows():
                # تحويل القيم لنصوص والتأكد إنها مش فاضية
                atype = str(row['Type']) if pd.notnull(row['Type']) else ""
                atime = str(row['Time']) if pd.notnull(row['Time']) else ""
                aperiod = str(row['Period']).upper() if pd.notnull(row['Period']) else ""
                
                # 🔴 قراءة اليوم
                aday = str(row['Day']) if 'Day' in df.columns and pd.notnull(row['Day']) else "يومياً"
                
                if atype and atime:
                    p_ar = "صباحاً" if "AM" in aperiod else "مساءً"
                    # 🔴 نطق الموعد باليوم
                    schedule_text += f"{atype} {aday} الساعة {atime} {p_ar}، "
            
            speak(schedule_text)
        else:
            speak("ليس لديك مواعيد مسجلة.")
            
    except Exception as e:
        print(f"❌ Error displaying: {e}")
        # لو حصل مشكلة في العناوين، بنطبعها عشان نعرف العيب فين
        print("Current Columns found:", df.columns.tolist()) 
        speak("حدث خطأ أثناء قراءة جدول المواعيد")

def cancel_appointment(user_phrase):
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        best, row_idx = 0, None
        for r in range(2, ws.max_row + 1):
            val = str(ws.cell(r, 1).value or "")
            sc = fuzz.token_sort_ratio(user_phrase, val)
            if sc > best:
                best, row_idx = sc, r
        if row_idx and best >= 90:
            ws.delete_rows(row_idx)
            wb.save(EXCEL_PATH)
            return "تم إلغاء الموعد بنجاح."
        return "لم يتم العثور على الموعد المطلوب."
    except Exception:
        return "حدث خطأ أثناء الإلغاء."

def get_current_alarms(file_path):
    all_alarms = []
    if not os.path.exists(file_path): return all_alarms
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, 1).value
            t, ampm = ws.cell(row, 2).value, ws.cell(row, 3).value
            
            # 🔴 التعديل هنا: جلب اليوم من العمود الرابع
            day = ws.cell(row, 4).value or "يومياً"

            if t is not None and ampm is not None and name is not None:
                t_str = str(t).strip()
                if ":" in t_str:
                    h, m = map(int, t_str.split(":"))
                else:
                    h, m = int(float(t_str)), 0
                ampm_normalized = str(ampm).strip().lower()
                is_pm = "pm" in ampm_normalized or "مساء" in ampm_normalized
                is_am = "am" in ampm_normalized or "صباح" in ampm_normalized
                if is_pm and h < 12: h += 12
                elif is_am and h == 12: h = 0
                
                # 🔴 التعديل هنا: إرسال اليوم مع باقي البيانات
                all_alarms.append((h, m, str(name), str(day)))
    except: pass
    return all_alarms